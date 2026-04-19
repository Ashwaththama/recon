import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILL_HEADER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_HR_ONLY = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_PAYROLL_ONLY = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
FILL_MISMATCH = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FONT_BOLD = Font(bold=True)


def _write_df_to_sheet(ws, df: pd.DataFrame, row_fill: PatternFill) -> None:
    headers = list(df.columns)
    ws.append(headers)

    header_row = ws[1]
    for cell in header_row:
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for _, row in df.iterrows():
        ws.append([str(v) if pd.isna(v) == False and not isinstance(v, str) else ("" if pd.isna(v) else v) for v in row])
        for cell in ws[ws.max_row]:
            cell.fill = row_fill

    _auto_size_columns(ws)


def _auto_size_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_summary_sheet(ws, summary: dict) -> None:
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for metric, value in summary.items():
        ws.append([metric, value])
        ws[ws.max_row][0].font = FONT_BOLD

    _auto_size_columns(ws)


def build_workbook(result: dict) -> io.BytesIO:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, result["summary"])

    ws_hr = wb.create_sheet("HR Only")
    if not result["hr_only"].empty:
        _write_df_to_sheet(ws_hr, result["hr_only"], FILL_HR_ONLY)
    else:
        ws_hr.append(["No discrepancies of this type found."])

    ws_pay = wb.create_sheet("Payroll Only")
    if not result["payroll_only"].empty:
        _write_df_to_sheet(ws_pay, result["payroll_only"], FILL_PAYROLL_ONLY)
    else:
        ws_pay.append(["No discrepancies of this type found."])

    ws_mis = wb.create_sheet("Field Mismatches")
    if not result["mismatches"].empty:
        _write_df_to_sheet(ws_mis, result["mismatches"], FILL_MISMATCH)
    else:
        ws_mis.append(["No field-level discrepancies found."])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
